import calendar
from datetime import datetime
from io import BytesIO
from typing import List, Optional

import openpyxl
from aiogram import Router
from aiogram.filters import Command
from aiogram.types import Message, BufferedInputFile
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.chart import LineChart, BarChart, Reference
from sqlalchemy import select, func

from bot import sql
from config import ADMIN_IDS
from logging_config import logger
from config_bd.models import AsyncSessionLocal, Users, Payments, PaymentsStars, PaymentsCryptobot

router = Router()

REF_ZALIV = [
    '1012882762', '1751833324', '7715104509', '6045891248', '778794666',
    '6803123509', '7623377322', '8036879919', '8185054692', '7208737418',
    '7545883972', '7801801881', '7231201607', '7863386911', '7251811519',
    '7717099908', '6514719405', '8154969535', '8196772935', '7985311643',
    '7607443801', '7617180616', '7780587251', '7999153238', '8075803624',
    '7774377890', '7939767168'
]

EXCLUDE_IDS = list(range(45, 1046))


# ---------- Вспомогательные функции конвертации ----------
def convert_stars_to_rub(amount: int) -> Optional[int]:
    mapping = {
        66: 99,
        179: 269,
        199: 299,
        333: 499
    }
    return mapping.get(amount)


def convert_crypto_to_rub(currency: str, amount: str) -> Optional[int]:
    mapping = {
        'TON': {'0.9': 99, '2.5': 269, '2.8': 299, '4.6': 499},
        'USDT': {'1.3': 99, '3.5': 269, '4.0': 299, '6.5': 499}
    }
    return mapping.get(currency, {}).get(amount)


class PaymentRecord:
    """Унифицированная запись о платеже."""
    def __init__(self, amount: int, is_gift: bool, time_created: datetime):
        self.amount = amount
        self.is_gift = is_gift
        self.time_created = time_created


@router.message(Command(commands=['stat']))
async def stat_command(message: Message):
    """Статистика по пользователям с указанным Ref или stamp (только для админов)."""
    if message.from_user.id not in ADMIN_IDS:
        return

    args = message.text.split()
    if len(args) < 2:
        await message.answer("❌ Использование: /stat <аргумент>")
        return

    arg = args[1].strip()
    total, with_sub, with_tarif, total_payments, source = await sql.get_stat_by_ref_or_stamp(arg)

    if total is None:
        await message.answer(f"{arg} - нет совпадений")
    else:
        await message.answer(f"{arg} {total} {with_sub} {with_tarif} {total_payments}")


@router.message(Command(commands=['anal']))
async def analytics_command(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    now = datetime.now()
    start_date = datetime(now.year, now.month, 1, 0, 0, 0)
    last_day = calendar.monthrange(now.year, now.month)[1]
    end_date = datetime(now.year, now.month, last_day, 23, 59, 59)

    async with AsyncSessionLocal() as session:
        # --- 1. Новые пользователи, взяли ключ, подключились ---
        stmt_users = select(Users).where(
            Users.create_user.between(start_date, end_date),
            ~Users.id.in_(EXCLUDE_IDS)
        )
        result = await session.execute(stmt_users)
        users_data = result.scalars().all()

        new_total = []
        new_zaliv = []
        new_saraf = []
        key_total = []
        key_zaliv = []
        key_saraf = []
        connect_total = []
        connect_zaliv = []
        connect_saraf = []

        set_new_total = set()
        set_new_zaliv = set()
        set_new_saraf = set()

        for user in users_data:
            is_zaliv = (user.stamp != '') or (str(user.ref) in REF_ZALIV)
            uid = user.user_id

            new_total.append(uid)
            set_new_total.add(uid)
            if is_zaliv:
                new_zaliv.append(uid)
                set_new_zaliv.add(uid)
            else:
                new_saraf.append(uid)
                set_new_saraf.add(uid)

            if user.is_pay_null:
                key_total.append(uid)
                if is_zaliv:
                    key_zaliv.append(uid)
                else:
                    key_saraf.append(uid)

            if user.is_tarif:
                connect_total.append(uid)
                if is_zaliv:
                    connect_zaliv.append(uid)
                else:
                    connect_saraf.append(uid)

        # --- 2. Множество пользователей, которые когда-либо платили ---
        # Основные платежи
        stmt_main = select(Payments.user_id).distinct().where(
            Payments.status == 'confirmed',
            Payments.amount != 1
        )
        result_main = await session.execute(stmt_main)
        paid_main = {row[0] for row in result_main.all()}

        # Звёздные платежи
        stmt_stars = select(PaymentsStars.user_id).distinct().where(
            PaymentsStars.status == 'confirmed'
        )
        result_stars = await session.execute(stmt_stars)
        paid_stars = {row[0] for row in result_stars.all()}

        # Крипто-платежи
        stmt_crypto = select(PaymentsCryptobot.user_id).distinct().where(
            PaymentsCryptobot.status == 'paid',
            PaymentsCryptobot.amount > 0.02
        )
        result_crypto = await session.execute(stmt_crypto)
        paid_crypto = {row[0] for row in result_crypto.all()}

        all_paid_users = paid_main.union(paid_stars).union(paid_crypto)

        # --- 3. Платежи новых пользователей за период (конвертированные) ---
        # Собираем все платежи за период
        all_period_payments: List[PaymentRecord] = []

        # Основные
        stmt_main_period = select(
            Payments.user_id,
            Payments.amount,
            Payments.is_gift,
            Payments.time_created
        ).where(
            Payments.time_created.between(start_date, end_date),
            Payments.amount != 1,
            Payments.status == 'confirmed'
        )
        result_main_period = await session.execute(stmt_main_period)
        for user_id, amount, is_gift, time_created in result_main_period.all():
            all_period_payments.append(PaymentRecord(amount, is_gift, time_created))

        # Звёздные
        stmt_stars_period = select(
            PaymentsStars.user_id,
            PaymentsStars.amount,
            PaymentsStars.is_gift,
            PaymentsStars.time_created
        ).where(
            PaymentsStars.time_created.between(start_date, end_date),
            PaymentsStars.status == 'confirmed'
        )
        result_stars_period = await session.execute(stmt_stars_period)
        for user_id, amount, is_gift, time_created in result_stars_period.all():
            rub = convert_stars_to_rub(amount)
            if rub:
                all_period_payments.append(PaymentRecord(rub, is_gift, time_created))

        # Крипто
        stmt_crypto_period = select(
            PaymentsCryptobot.user_id,
            PaymentsCryptobot.amount,
            PaymentsCryptobot.currency,
            PaymentsCryptobot.is_gift,
            PaymentsCryptobot.time_created
        ).where(
            PaymentsCryptobot.time_created.between(start_date, end_date),
            PaymentsCryptobot.status == 'paid',
            PaymentsCryptobot.amount > 0.02
        )
        result_crypto_period = await session.execute(stmt_crypto_period)
        for user_id, amount, currency, is_gift, time_created in result_crypto_period.all():
            rub = convert_crypto_to_rub(currency, str(amount))
            if rub:
                all_period_payments.append(PaymentRecord(rub, is_gift, time_created))

        # Платежи новых пользователей
        new_payments_data = []  # (user_id, amount)

        # Основные
        stmt_main_new = select(Payments.user_id, Payments.amount).where(
            Payments.time_created.between(start_date, end_date),
            Payments.amount != 1,
            Payments.status == 'confirmed'
        )
        result_main_new = await session.execute(stmt_main_new)
        for uid, amt in result_main_new.all():
            if uid in set_new_total:
                new_payments_data.append((uid, amt))

        # Звёзды
        stmt_stars_new = select(PaymentsStars.user_id, PaymentsStars.amount).where(
            PaymentsStars.time_created.between(start_date, end_date),
            PaymentsStars.status == 'confirmed'
        )
        result_stars_new = await session.execute(stmt_stars_new)
        for uid, amt in result_stars_new.all():
            if uid in set_new_total:
                rub = convert_stars_to_rub(amt)
                if rub:
                    new_payments_data.append((uid, rub))

        # Крипто
        stmt_crypto_new = select(
            PaymentsCryptobot.user_id,
            PaymentsCryptobot.amount,
            PaymentsCryptobot.currency
        ).where(
            PaymentsCryptobot.time_created.between(start_date, end_date),
            PaymentsCryptobot.status == 'paid',
            PaymentsCryptobot.amount > 0.02
        )
        result_crypto_new = await session.execute(stmt_crypto_new)
        for uid, amt, cur in result_crypto_new.all():
            if uid in set_new_total:
                rub = convert_crypto_to_rub(cur, str(amt))
                if rub:
                    new_payments_data.append((uid, rub))

        pay_sum_total = 0
        pay_sum_zaliv = 0
        pay_sum_saraf = 0
        pay_users_total = set()
        pay_users_zaliv = set()
        pay_users_saraf = set()

        for uid, amount in new_payments_data:
            pay_sum_total += amount
            pay_users_total.add(uid)
            if uid in set_new_zaliv:
                pay_sum_zaliv += amount
                pay_users_zaliv.add(uid)
            elif uid in set_new_saraf:
                pay_sum_saraf += amount
                pay_users_saraf.add(uid)

        # --- 4. Общая статистика всех платежей за период ---
        all_payments = []  # (amount, is_gift, time_created)

        # Основные
        stmt_main_all = select(Payments.amount, Payments.is_gift, Payments.time_created).where(
            Payments.time_created.between(start_date, end_date),
            Payments.amount != 1,
            Payments.status == 'confirmed'
        )
        result_main_all = await session.execute(stmt_main_all)
        for amount, is_gift, time_created in result_main_all.all():
            all_payments.append((amount, is_gift, time_created))

        # Звёзды
        stmt_stars_all = select(PaymentsStars.amount, PaymentsStars.is_gift, PaymentsStars.time_created).where(
            PaymentsStars.time_created.between(start_date, end_date),
            PaymentsStars.status == 'confirmed'
        )
        result_stars_all = await session.execute(stmt_stars_all)
        for amount, is_gift, time_created in result_stars_all.all():
            rub = convert_stars_to_rub(amount)
            if rub:
                all_payments.append((rub, is_gift, time_created))

        # Крипто
        stmt_crypto_all = select(
            PaymentsCryptobot.amount,
            PaymentsCryptobot.currency,
            PaymentsCryptobot.is_gift,
            PaymentsCryptobot.time_created
        ).where(
            PaymentsCryptobot.time_created.between(start_date, end_date),
            PaymentsCryptobot.status == 'paid',
            PaymentsCryptobot.amount > 0.02
        )
        result_crypto_all = await session.execute(stmt_crypto_all)
        for amount, currency, is_gift, time_created in result_crypto_all.all():
            rub = convert_crypto_to_rub(currency, str(amount))
            if rub:
                all_payments.append((rub, is_gift, time_created))

        total_revenue = sum(p[0] for p in all_payments)
        total_payments_count = len(all_payments)
        aov = total_revenue / total_payments_count if total_payments_count else 0

        stmt_total_users = select(func.count(Users.id)).where(~Users.id.in_(EXCLUDE_IDS))
        total_users_count = (await session.execute(stmt_total_users)).scalar() or 0
        arpu = total_revenue / total_users_count if total_users_count else 0

        # Разбивка по суммам
        sum_99_count = sum_99_amount = 0
        sum_269_count = sum_269_amount = 0
        sum_299_count = sum_299_amount = 0
        sum_499_count = sum_499_amount = 0
        gift_count = gift_amount = 0

        for amount, is_gift, _ in all_payments:
            if is_gift:
                gift_count += 1
                gift_amount += amount
            else:
                if amount == 99:
                    sum_99_count += 1
                    sum_99_amount += amount
                elif amount == 269:
                    sum_269_count += 1
                    sum_269_amount += amount
                elif amount == 299:
                    sum_299_count += 1
                    sum_299_amount += amount
                elif amount == 499:
                    sum_499_count += 1
                    sum_499_amount += amount

        # Разбивка по 4 периодам внутри месяца
        total_days = last_day
        chunk_size = total_days // 4
        period_starts = []
        period_ends = []
        for i in range(4):
            start_day = 1 + i * chunk_size
            end_day = start_day + chunk_size - 1 if i < 3 else last_day
            period_starts.append(datetime(now.year, now.month, start_day, 0, 0, 0))
            period_ends.append(datetime(now.year, now.month, end_day, 23, 59, 59))

        period_revenues = [0, 0, 0, 0]
        period_counts = [0, 0, 0, 0]

        for amount, is_gift, time_created in all_payments:
            for i, (p_start, p_end) in enumerate(zip(period_starts, period_ends)):
                if p_start <= time_created <= p_end:
                    period_revenues[i] += amount
                    period_counts[i] += 1
                    break

        period_lines = []
        for i in range(4):
            rev = period_revenues[i]
            cnt = period_counts[i]
            avg = rev / cnt if cnt else 0
            period_lines.append(
                f"{i+1} Период ({period_starts[i].strftime('%d.%m')} – {period_ends[i].strftime('%d.%m')}): "
                f"{rev} ₽ / {cnt} плат. (ср. {avg:.2f} ₽)"
            )

    # --- Формирование отчёта ---
    report = (
        f"📊 Аналитика за период {start_date.strftime('%d.%m.%Y')} – {end_date.strftime('%d.%m.%Y')}\n\n"
        f"👥 <b>Новые пользователи:</b>\n"
        f"├ Всего: {len(new_total)}\n"
        f"├ Залив: {len(new_zaliv)}\n"
        f"└ Сарафанка: {len(new_saraf)}\n\n"
        f"🔑 <b>Взяли ключ:</b>\n"
        f"├ Всего: {len(key_total)}\n"
        f"├ Залив: {len(key_zaliv)}\n"
        f"└ Сарафанка: {len(key_saraf)}\n\n"
        f"🔗 <b>Подключились:</b>\n"
        f"├ Всего: {len(connect_total)}\n"
        f"├ Залив: {len(connect_zaliv)}\n"
        f"└ Сарафанка: {len(connect_saraf)}\n\n"
        f"💰 <b>Платежи новых пользователей (сумма, исключая 1₽):</b>\n"
        f"├ Всего: {pay_sum_total} ₽ (уникальных плательщиков: {len(pay_users_total)})\n"
        f"├ Залив: {pay_sum_zaliv} ₽ (уникальных: {len(pay_users_zaliv)})\n"
        f"└ Сарафанка: {pay_sum_saraf} ₽ (уникальных: {len(pay_users_saraf)})\n\n"
        f"📈 <b>Общая статистика платежей за период:</b>\n"
        f"├ Общая выручка: {total_revenue} ₽\n"
        f"├ Всего платежей: {total_payments_count}\n"
        f"├ AOV (средний чек): {aov:.2f} ₽\n"
        f"├ ARPU (на всех пользователей*): {arpu:.2f} ₽\n"
        f"├ Платежей 99₽: {sum_99_count} шт., сумма {sum_99_amount} ₽\n"
        f"├ Платежей 269₽: {sum_269_count} шт., сумма {sum_269_amount} ₽\n"
        f"├ Платежей 299₽: {sum_299_count} шт., сумма {sum_299_amount} ₽\n"
        f"├ Платежей 499₽: {sum_499_count} шт., сумма {sum_499_amount} ₽\n"
        f"└ Подарки (is_gift): {gift_count} шт., сумма {gift_amount} ₽\n\n"
        f"📅 <b>Доход по периодам:</b>\n"
    )
    for line in period_lines:
        report += f"├ {line}\n"
    report += f"\n* – общее количество пользователей (исключая ID 45–1045): {total_users_count}"

    await message.answer(report)


@router.message(Command(commands=['anal_export']))
async def analytics_export(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ Команда доступна только администраторам.")
        return

    await message.answer("🔄 Формирую помесячную аналитику...")

    try:
        now = datetime.now()
        current_year = now.year
        current_month = now.month

        months = [(current_year, month) for month in range(1, current_month + 1)]

        monthly_data = {}
        daily_data_by_month = {}

        for year, month in months:
            start_date = datetime(year, month, 1, 0, 0, 0)
            last_day = calendar.monthrange(year, month)[1]
            end_date = datetime(year, month, last_day, 23, 59, 59)
            month_key = start_date.strftime('%B %Y')

            async with AsyncSessionLocal() as session:
                # --- Новые пользователи за месяц ---
                stmt_new_users = select(Users).where(
                    Users.create_user.between(start_date, end_date),
                    ~Users.id.in_(EXCLUDE_IDS)
                )
                result = await session.execute(stmt_new_users)
                new_users = result.scalars().all()

                new_total = []
                new_zaliv = []
                new_saraf = []
                key_total = []
                key_zaliv = []
                key_saraf = []
                connect_total = []
                connect_zaliv = []
                connect_saraf = []
                set_new_total = set()
                set_new_zaliv = set()
                set_new_saraf = set()

                daily_stats = {day: {'new': 0, 'key': 0, 'connect': 0, 'paid': 0} for day in range(1, last_day + 1)}

                for user in new_users:
                    is_zaliv = (user.stamp != '') or (str(user.ref) in REF_ZALIV)
                    uid = user.user_id
                    create_day = user.create_user.day

                    new_total.append(uid)
                    set_new_total.add(uid)
                    if is_zaliv:
                        new_zaliv.append(uid)
                        set_new_zaliv.add(uid)
                    else:
                        new_saraf.append(uid)
                        set_new_saraf.add(uid)

                    if user.is_pay_null:
                        key_total.append(uid)
                        if is_zaliv:
                            key_zaliv.append(uid)
                        else:
                            key_saraf.append(uid)

                    if user.is_tarif:
                        connect_total.append(uid)
                        if is_zaliv:
                            connect_zaliv.append(uid)
                        else:
                            connect_saraf.append(uid)

                    daily_stats[create_day]['new'] += 1
                    if user.is_pay_null:
                        daily_stats[create_day]['key'] += 1
                    if user.is_tarif:
                        daily_stats[create_day]['connect'] += 1

                # --- Множество плативших ---
                stmt_paid_main = select(Payments.user_id).distinct().where(
                    Payments.status == 'confirmed',
                    Payments.amount != 1
                )
                paid_main = {row[0] for row in (await session.execute(stmt_paid_main)).all()}

                stmt_paid_stars = select(PaymentsStars.user_id).distinct().where(
                    PaymentsStars.status == 'confirmed'
                )
                paid_stars = {row[0] for row in (await session.execute(stmt_paid_stars)).all()}

                stmt_paid_crypto = select(PaymentsCryptobot.user_id).distinct().where(
                    PaymentsCryptobot.status == 'paid',
                    PaymentsCryptobot.amount > 0.02
                )
                paid_crypto = {row[0] for row in (await session.execute(stmt_paid_crypto)).all()}

                all_paid_users = paid_main.union(paid_stars).union(paid_crypto)

                for uid in set_new_total:
                    if uid in all_paid_users:
                        # найдём день регистрации
                        for user in new_users:
                            if user.user_id == uid:
                                daily_stats[user.create_user.day]['paid'] += 1
                                break

                # --- Платежи новых пользователей за этот месяц ---
                new_payments_amounts = []

                # Основные
                stmt_main_new = select(Payments.user_id, Payments.amount).where(
                    Payments.time_created.between(start_date, end_date),
                    Payments.amount != 1,
                    Payments.status == 'confirmed'
                )
                for uid, amt in (await session.execute(stmt_main_new)).all():
                    if uid in set_new_total:
                        new_payments_amounts.append((uid, amt))

                # Звёзды
                stmt_stars_new = select(PaymentsStars.user_id, PaymentsStars.amount).where(
                    PaymentsStars.time_created.between(start_date, end_date),
                    PaymentsStars.status == 'confirmed'
                )
                for uid, amt in (await session.execute(stmt_stars_new)).all():
                    if uid in set_new_total:
                        rub = convert_stars_to_rub(amt)
                        if rub:
                            new_payments_amounts.append((uid, rub))

                # Крипто
                stmt_crypto_new = select(
                    PaymentsCryptobot.user_id,
                    PaymentsCryptobot.amount,
                    PaymentsCryptobot.currency
                ).where(
                    PaymentsCryptobot.time_created.between(start_date, end_date),
                    PaymentsCryptobot.status == 'paid',
                    PaymentsCryptobot.amount > 0.02
                )
                for uid, amt, cur in (await session.execute(stmt_crypto_new)).all():
                    if uid in set_new_total:
                        rub = convert_crypto_to_rub(cur, str(amt))
                        if rub:
                            new_payments_amounts.append((uid, rub))

                pay_sum_total = 0
                pay_sum_zaliv = 0
                pay_sum_saraf = 0
                pay_users_total = set()
                pay_users_zaliv = set()
                pay_users_saraf = set()

                for uid, amount in new_payments_amounts:
                    pay_sum_total += amount
                    pay_users_total.add(uid)
                    if uid in set_new_zaliv:
                        pay_sum_zaliv += amount
                        pay_users_zaliv.add(uid)
                    elif uid in set_new_saraf:
                        pay_sum_saraf += amount
                        pay_users_saraf.add(uid)

                # --- Общие платежи за месяц (все пользователи) ---
                all_payments = []  # (amount, is_gift)

                # Основные
                stmt_main_all = select(Payments.amount, Payments.is_gift).where(
                    Payments.time_created.between(start_date, end_date),
                    Payments.amount != 1,
                    Payments.status == 'confirmed'
                )
                for amount, is_gift in (await session.execute(stmt_main_all)).all():
                    all_payments.append((amount, is_gift))

                # Звёзды
                stmt_stars_all = select(PaymentsStars.amount, PaymentsStars.is_gift).where(
                    PaymentsStars.time_created.between(start_date, end_date),
                    PaymentsStars.status == 'confirmed'
                )
                for amount, is_gift in (await session.execute(stmt_stars_all)).all():
                    rub = convert_stars_to_rub(amount)
                    if rub:
                        all_payments.append((rub, is_gift))

                # Крипто
                stmt_crypto_all = select(
                    PaymentsCryptobot.amount,
                    PaymentsCryptobot.currency,
                    PaymentsCryptobot.is_gift
                ).where(
                    PaymentsCryptobot.time_created.between(start_date, end_date),
                    PaymentsCryptobot.status == 'paid',
                    PaymentsCryptobot.amount > 0.02
                )
                for amount, currency, is_gift in (await session.execute(stmt_crypto_all)).all():
                    rub = convert_crypto_to_rub(currency, str(amount))
                    if rub:
                        all_payments.append((rub, is_gift))

                total_revenue = sum(p[0] for p in all_payments)
                total_payments_count = len(all_payments)
                aov = total_revenue / total_payments_count if total_payments_count else 0

                stmt_cumulative = select(func.count(Users.id)).where(
                    Users.create_user <= end_date,
                    ~Users.id.in_(EXCLUDE_IDS)
                )
                cumulative_users = (await session.execute(stmt_cumulative)).scalar() or 1
                arpu = total_revenue / cumulative_users

                # Разбивка по суммам
                sum_99_count = sum_99_amount = 0
                sum_269_count = sum_269_amount = 0
                sum_299_count = sum_299_amount = 0
                sum_499_count = sum_499_amount = 0
                gift_count = gift_amount = 0

                for amount, is_gift in all_payments:
                    if is_gift:
                        gift_count += 1
                        gift_amount += amount
                    else:
                        if amount == 99:
                            sum_99_count += 1
                            sum_99_amount += amount
                        elif amount == 269:
                            sum_269_count += 1
                            sum_269_amount += amount
                        elif amount == 299:
                            sum_299_count += 1
                            sum_299_amount += amount
                        elif amount == 499:
                            sum_499_count += 1
                            sum_499_amount += amount

                monthly_data[month_key] = {
                    'new_total': len(new_total),
                    'new_zaliv': len(new_zaliv),
                    'new_saraf': len(new_saraf),
                    'key_total': len(key_total),
                    'key_zaliv': len(key_zaliv),
                    'key_saraf': len(key_saraf),
                    'connect_total': len(connect_total),
                    'connect_zaliv': len(connect_zaliv),
                    'connect_saraf': len(connect_saraf),
                    'pay_new_sum_total': pay_sum_total,
                    'pay_new_users_total': len(pay_users_total),
                    'pay_new_sum_zaliv': pay_sum_zaliv,
                    'pay_new_users_zaliv': len(pay_users_zaliv),
                    'pay_new_sum_saraf': pay_sum_saraf,
                    'pay_new_users_saraf': len(pay_users_saraf),
                    'total_revenue': total_revenue,
                    'total_payments': total_payments_count,
                    'aov': aov,
                    'arpu': arpu,
                    'cumulative_users': cumulative_users,
                    'sum_99_count': sum_99_count,
                    'sum_99_amount': sum_99_amount,
                    'sum_269_count': sum_269_count,
                    'sum_269_amount': sum_269_amount,
                    'sum_299_count': sum_299_count,
                    'sum_299_amount': sum_299_amount,
                    'sum_499_count': sum_499_count,
                    'sum_499_amount': sum_499_amount,
                    'gift_count': gift_count,
                    'gift_amount': gift_amount,
                }

                # --- Поденные данные (кумулятивные) ---
                stmt_before = select(Users.user_id, Users.is_pay_null, Users.is_tarif).where(
                    Users.create_user < start_date,
                    ~Users.id.in_(EXCLUDE_IDS)
                )
                users_before = (await session.execute(stmt_before)).all()
                cum_users_before = len(users_before)
                cum_key_before = sum(1 for u in users_before if u.is_pay_null)
                cum_connect_before = sum(1 for u in users_before if u.is_tarif)

                daily_cumulative = []
                cum_users = cum_users_before
                cum_key = cum_key_before
                cum_connect = cum_connect_before

                for day in range(1, last_day + 1):
                    cum_users += daily_stats[day]['new']
                    cum_key += daily_stats[day]['key']
                    cum_connect += daily_stats[day]['connect']
                    daily_cumulative.append({
                        'day': day,
                        'cum_users': cum_users,
                        'cum_key': cum_key,
                        'cum_connect': cum_connect,
                        'new': daily_stats[day]['new'],
                        'key': daily_stats[day]['key'],
                        'connect': daily_stats[day]['connect'],
                        'paid': daily_stats[day]['paid']
                    })

                daily_data_by_month[month_key] = daily_cumulative

        # --- Создание Excel файла ---
        wb = openpyxl.Workbook()
        ws_main = wb.active
        ws_main.title = "Помесячная аналитика"

        headers = ['Показатель'] + list(monthly_data.keys())
        ws_main.append(headers)

        metric_rows = [
            ('Новые пользователи (всего)', 'new_total'),
            ('Новые пользователи (залив)', 'new_zaliv'),
            ('Новые пользователи (сарафан)', 'new_saraf'),
            ('Взяли ключ (всего)', 'key_total'),
            ('Взяли ключ (залив)', 'key_zaliv'),
            ('Взяли ключ (сарафан)', 'key_saraf'),
            ('Подключились (всего)', 'connect_total'),
            ('Подключились (залив)', 'connect_zaliv'),
            ('Подключились (сарафан)', 'connect_saraf'),
            ('Платежи новых (сумма, всего)', 'pay_new_sum_total'),
            ('Платежи новых (уникальных, всего)', 'pay_new_users_total'),
            ('Платежи новых (сумма, залив)', 'pay_new_sum_zaliv'),
            ('Платежи новых (уникальных, залив)', 'pay_new_users_zaliv'),
            ('Платежи новых (сумма, сарафан)', 'pay_new_sum_saraf'),
            ('Платежи новых (уникальных, сарафан)', 'pay_new_users_saraf'),
            ('Общая выручка (₽)', 'total_revenue'),
            ('Количество платежей', 'total_payments'),
            ('AOV (₽)', 'aov'),
            ('ARPU (₽)', 'arpu'),
            ('Пользователей на конец месяца', 'cumulative_users'),
            ('Платежей 99₽ (шт)', 'sum_99_count'),
            ('Сумма 99₽ (₽)', 'sum_99_amount'),
            ('Платежей 269₽ (шт)', 'sum_269_count'),
            ('Сумма 269₽ (₽)', 'sum_269_amount'),
            ('Платежей 299₽ (шт)', 'sum_299_count'),
            ('Сумма 299₽ (₽)', 'sum_299_amount'),
            ('Платежей 499₽ (шт)', 'sum_499_count'),
            ('Сумма 499₽ (₽)', 'sum_499_amount'),
            ('Подарков (шт)', 'gift_count'),
            ('Сумма подарков (₽)', 'gift_amount'),
        ]

        row_idx = 2
        for label, key in metric_rows:
            row = [label]
            ws_main.append(row)
            col_idx = 2
            for month in monthly_data.keys():
                value = monthly_data[month].get(key, 0)
                if key in ('aov', 'arpu'):
                    cell_value = round(value, 2)
                else:
                    cell_value = value if isinstance(value, int) else round(value, 2)
                ws_main.cell(row=row_idx, column=col_idx, value=cell_value)
                col_idx += 1
            row_idx += 1

        # Оформление
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws_main[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        month_columns = list(monthly_data.keys())
        for r in range(2, row_idx):
            for c in range(1, ws_main.max_column + 1):
                ws_main.cell(row=r, column=c).border = thin_border
            jan_cell = ws_main.cell(row=r, column=2)
            jan_cell.fill = yellow_fill
            for col_idx in range(3, 2 + len(month_columns)):
                current = ws_main.cell(row=r, column=col_idx)
                prev = ws_main.cell(row=r, column=col_idx-1)
                try:
                    cur_val = float(current.value)
                    prev_val = float(prev.value)
                except (TypeError, ValueError):
                    continue
                if cur_val > prev_val:
                    current.fill = light_green_fill
                elif cur_val < prev_val:
                    current.fill = light_red_fill

        for col in ws_main.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_main.column_dimensions[col_letter].width = min(max_len + 2, 50)

        ws_main.freeze_panes = 'B2'

        # Листы по месяцам с графиками
        for month_key, daily_data in daily_data_by_month.items():
            ws = wb.create_sheet(title=month_key[:31])
            ws.append(['День', 'Новые', 'Взяли ключ', 'Подключились', 'Платили',
                       'Всего пользователей (накопительно)', 'Всего ключей (накопительно)', 'Всего подключений (накопительно)'])
            for d in daily_data:
                ws.append([
                    d['day'],
                    d['new'],
                    d['key'],
                    d['connect'],
                    d['paid'],
                    d['cum_users'],
                    d['cum_key'],
                    d['cum_connect']
                ])

            for row in ws.iter_rows(min_row=1, max_row=len(daily_data)+1, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 20)

            # Линейный график (накопительные)
            chart1 = LineChart()
            chart1.title = "Накопительные показатели"
            chart1.style = 13
            chart1.y_axis.title = "Количество"
            chart1.x_axis.title = "День месяца"
            data = Reference(ws, min_col=6, max_col=8, min_row=1, max_row=len(daily_data)+1)
            dates = Reference(ws, min_col=1, min_row=2, max_row=len(daily_data)+1)
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(dates)
            if len(chart1.series) >= 3:
                chart1.series[0].graphicalProperties.line.solidFill = "0000FF"
                chart1.series[1].graphicalProperties.line.solidFill = "00B0F0"
                chart1.series[2].graphicalProperties.line.solidFill = "000000"
            ws.add_chart(chart1, "J2")

            # Столбцовая диаграмма (ежедневные)
            chart2 = BarChart()
            chart2.title = "Ежедневные показатели"
            chart2.style = 13
            chart2.y_axis.title = "Количество"
            chart2.x_axis.title = "День месяца"
            data2 = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=len(daily_data)+1)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(dates)
            ws.add_chart(chart2, "J20")

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        await message.answer_document(
            document=BufferedInputFile(
                excel_file.read(),
                filename=f"analytics_{current_year}_{current_month}.xlsx"
            ),
            caption=f"📊 Помесячная аналитика с января {current_year} по {now.strftime('%B %Y')}"
        )

        logger.info(f"Админ {message.from_user.id} выгрузил помесячную аналитику")

    except Exception as e:
        logger.exception("Ошибка при экспорте помесячной аналитики")
        await message.answer(f"❌ Ошибка: {str(e)}")
