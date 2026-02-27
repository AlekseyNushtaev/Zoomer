from datetime import datetime
import openpyxl
from aiogram import Router
from openpyxl.styles import Alignment, Border, Side

from bot import sql
from config import ADMIN_IDS
from logging_config import logger
from aiogram.types import Message, FSInputFile
from aiogram.filters import Command

router = Router()


@router.message(Command(commands=['export']))
async def export_database_to_excel(message: Message):
    """Экспорт базы данных в Excel файл"""
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ Эта команда доступна только администраторам.")
        return

    try:
        await message.answer("🔄 Начинаю экспорт базы данных...")

        # Получаем все данные через асинхронные методы
        users_list = await sql.get_all_users()
        payments_list = await sql.get_all_payments()
        payments_stars_list = await sql.get_all_payments_stars()
        payments_cryptobot_list = await sql.get_all_payments_cryptobot()
        gifts_list = await sql.get_all_gifts()
        online_list = await sql.get_all_online()
        white_counter_list = await sql.get_all_white_counter()

        # Создаём книгу Excel
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # --- Лист USERS ---
        ws_users = wb.create_sheet(title="users")
        users_columns = [
            'ID', 'User ID', 'Ref', 'Is_delete', 'Is_pay_null', 'Is_tarif',
            'Create_user', 'Is_admin', 'has_discount', 'subscription_end_date',
            'white_subscription_end_date', 'last_notification_date',
            'last_broadcast_status', 'last_broadcast_date', 'stamp', 'ttclid'
        ]
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Заголовки
        for col_num, title in enumerate(users_columns, 1):
            cell = ws_users.cell(row=1, column=col_num, value=title)
            cell.alignment = header_alignment
            cell.border = thin_border

        # Данные
        for row_num, user in enumerate(users_list, 2):
            row_data = [
                user.id, user.user_id, user.ref, user.is_delete,
                user.is_pay_null, user.is_tarif, user.create_user,
                user.is_admin, user.has_discount, user.subscription_end_date,
                user.white_subscription_end_date, user.last_notification_date,
                user.last_broadcast_status, user.last_broadcast_date,
                user.stamp, user.ttclid
            ]
            for col_num, value in enumerate(row_data, 1):
                # Форматирование дат
                if col_num in (10, 11, 14) and value:  # subscription_end_date, white_subscription_end_date, last_broadcast_date
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif col_num == 12 and value:  # last_notification_date
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d')
                cell = ws_users.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        # Автоширина
        for col in ws_users.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_users.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # --- Лист PAYMENTS (Platega) ---
        ws_payments = wb.create_sheet(title="payments_sbp")
        payments_columns = ['ID', 'User ID', 'Amount', 'Time Created', 'Is Gift', 'Status', 'Transaction_Id']
        for col_num, title in enumerate(payments_columns, 1):
            cell = ws_payments.cell(row=1, column=col_num, value=title)
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, pay in enumerate(payments_list, 2):
            row_data = [
                pay.id, pay.user_id, pay.amount, pay.time_created,
                pay.is_gift, pay.status, pay.transaction_id
            ]
            for col_num, value in enumerate(row_data, 1):
                if col_num == 4 and value and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws_payments.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        for col in ws_payments.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_payments.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # --- Лист PAYMENTS_STARS ---
        ws_payments_stars = wb.create_sheet(title="payments_stars")
        stars_columns = ['ID', 'User ID', 'Amount (Stars)', 'Time Created', 'Is Gift', 'Status']
        for col_num, title in enumerate(stars_columns, 1):
            cell = ws_payments_stars.cell(row=1, column=col_num, value=title)
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, ps in enumerate(payments_stars_list, 2):
            row_data = [
                ps.id, ps.user_id, ps.amount, ps.time_created,
                ps.is_gift, ps.status
            ]
            for col_num, value in enumerate(row_data, 1):
                if col_num == 4 and value and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws_payments_stars.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        for col in ws_payments_stars.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_payments_stars.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # --- Лист PAYMENTS_CRYPTOBOT ---
        ws_payments_cryptobot = wb.create_sheet(title="payments_cryptobot")
        crypto_columns = [
            'ID', 'User ID', 'Amount', 'Currency', 'Time Created',
            'Is Gift', 'Status', 'Invoice ID', 'Payload'
        ]
        for col_num, title in enumerate(crypto_columns, 1):
            cell = ws_payments_cryptobot.cell(row=1, column=col_num, value=title)
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, pc in enumerate(payments_cryptobot_list, 2):
            row_data = [
                pc.id, pc.user_id, pc.amount, pc.currency, pc.time_created,
                pc.is_gift, pc.status, pc.invoice_id, pc.payload
            ]
            for col_num, value in enumerate(row_data, 1):
                if col_num == 5 and value and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws_payments_cryptobot.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        for col in ws_payments_cryptobot.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_payments_cryptobot.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # --- Лист GIFTS ---
        ws_gifts = wb.create_sheet(title="gifts")
        gifts_columns = ['gift_id', 'giver_id', 'duration', 'recepient_id', 'white_flag', 'flag']
        for col_num, title in enumerate(gifts_columns, 1):
            cell = ws_gifts.cell(row=1, column=col_num, value=title)
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, gift in enumerate(gifts_list, 2):
            row_data = [
                gift.gift_id, gift.giver_id, gift.duration,
                gift.recepient_id, gift.white_flag, gift.flag
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws_gifts.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        for col in ws_gifts.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_gifts.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # --- Лист ONLINE ---
        ws_online = wb.create_sheet(title="online")
        online_columns = ['ID', 'Дата сбора', 'Всего в панели', 'Активны сегодня', 'Платных', 'Триальных']
        for col_num, title in enumerate(online_columns, 1):
            cell = ws_online.cell(row=1, column=col_num, value=title)
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, rec in enumerate(online_list, 2):
            row_data = [
                rec.online_id, rec.online_date, rec.users_panel,
                rec.users_active, rec.users_pay, rec.users_trial
            ]
            for col_num, value in enumerate(row_data, 1):
                if col_num == 2 and value and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws_online.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        for col in ws_online.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_online.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # --- Лист WHITE_COUNTER ---
        ws_white_counter = wb.create_sheet(title="white_counter")
        wc_columns = ['ID', 'User ID', 'Time Created']
        for col_num, title in enumerate(wc_columns, 1):
            cell = ws_white_counter.cell(row=1, column=col_num, value=title)
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, wc in enumerate(white_counter_list, 2):
            row_data = [wc.id, wc.user_id, wc.time_created]
            for col_num, value in enumerate(row_data, 1):
                if col_num == 3 and value and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws_white_counter.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        for col in ws_white_counter.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_white_counter.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Заморозка заголовков
        for ws in [ws_users, ws_payments, ws_payments_stars, ws_payments_cryptobot,
                   ws_gifts, ws_online, ws_white_counter]:
            ws.freeze_panes = ws['A2']

        # Сохраняем файл
        wb.save('export.xlsx')

        # Статистика для caption
        users_count = len(users_list)
        gifts_count = len(gifts_list)
        payments_count = len(payments_list)
        payments_stars_count = len(payments_stars_list)
        payments_cryptobot_count = len(payments_cryptobot_list)
        white_counter_count = len(white_counter_list)

        white_subscription_count = sum(
            1 for u in users_list if u.white_subscription_end_date is not None
        )

        await message.answer_document(
            document=FSInputFile('export.xlsx'),
            caption=f"📊 Экспорт базы данных\n"
                    f"📅 Создано: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
                    f"📊 Статистика:\n"
                    f"├ 👥 Пользователей: {users_count}\n"
                    f"├ 🎁 Подарков: {gifts_count}\n"
                    f"├ 💰 Платежей Platega: {payments_count}\n"
                    f"├ ⭐ Платежей Stars: {payments_stars_count}\n"
                    f"├ 💎 Крипто-платежей: {payments_cryptobot_count}\n"
                    f"├ ⚪ White-подписок: {white_subscription_count}\n"
                    f"└ 👁 White-кликов: {white_counter_count}"
        )

        logger.info(f"Администратор {message.from_user.id} экспортировал базу данных в Excel")

    except Exception as e:
        error_message = f"❌ Ошибка при экспорте базы данных: {str(e)}"
        logger.error(error_message)
        logger.exception("Детали ошибки:")
        await message.answer(error_message)